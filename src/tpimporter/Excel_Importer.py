import traceback
import os
import sys
from openpyxl import load_workbook, Workbook
import msoffcrypto
from typing import List, Dict


class Excel_Importer:
    """
    basic class to transform excel files to a list of dictionaries;
    tp_map must be a dictionary with the following structure:
    {'fieldname': column-number (integer), ...}
    the 2 callback signals can be used to report the status of the process back to the user
    """

    callback_status_signal = None
    callback_progress_signal = None
    tp_map = dict()

    def __init__(self, valid_statuses=("ok", "change", "new")):
        """
        :param valid_statuses:
        """
        self.valid_statuses = valid_statuses

    def register_callback_signals(
        self, callback_status_signal=None, callback_progress_signal=None
    ):
        self.callback_status_signal = callback_status_signal
        self.callback_progress_signal = callback_progress_signal

    def callback_status(self, status: str):
        """
        a callback function that reports status updates - should be overwritten when subclassing this class
        :param status: str
        :return:
        """
        if self.callback_status_signal:
            try:
                self.callback_status_signal.emit(status)
            except:
                print(traceback.format_exc())
        else:
            print(status)

    def callback_progress(self, progress: int):
        """
        a callback function that reports progress updates - should be overwritten when subclassing this class
        :param progress: int
        :return:
        """
        if self.callback_progress_signal:
            try:
                self.callback_progress_signal.emit(progress)
            except:
                print(traceback.format_exc())
        else:
            print(progress)

    def get_tp_data_from_file(self, path) -> List[Dict]:
        """
        imports the Titelplanung xlsm from a given filepath (String or Path)
        :param path: filepath (String or Path)
        :return: a list of dicts (one dict per title)
        """
        wb = self._load_workbook(path)
        return self._get_data_from_wb(wb)

    def _load_workbook(self, path) -> Workbook:
        """
        loads an Excel Workbook from the given path
        :param path:
        :return:
        """
        try:
            self.callback_status("reading data from TP")
            self.callback_progress(0)
            wb = load_workbook(
                path, read_only=True, data_only=True, keep_vba=False, keep_links=False
            )
        except:
            # falls das Excel File passwortgeschützt ist schlägt der erste Versuch fehl;
            # dann muss erst das Passwort entfernt werden
            TEMP_NAME = self._getTempDirName(tempFileName="temp_tp.xlsm")
            self._removePasswordFromExcelFile(path, "uf", TEMP_NAME)
            self.callback_status("reading data from TP")
            wb = load_workbook(TEMP_NAME, read_only=True, data_only=True)
        self.callback_progress(20)
        return wb

    def _get_data_from_wb(
        self,
        workbook: Workbook,
        first_row: int = 8,
        worksheet_name: str = "TP",
        channel_type="transactional",
    ) -> List[Dict]:
        """
        fetches data from the given Excel Workbook instance
        :param workbook: a Workbook instance
        :param first_row: the first row where we should expect the data to begin
        :param worksheet_name: name of the worksheet that should be read
        :param channel_type: possible values: transactional/filmtastic/arthousecnma/homeofhorror
        :return:
        """
        tp_data = []
        ws = workbook[worksheet_name]
        i = first_row
        for row in ws["A" + str(first_row) : "QF10000"]:
            try:
                i += 1
                if (
                    row[self.tp_map.get("tnr")].value
                    and row[self.tp_map.get("status")].value in self.valid_statuses
                ):
                    row_data = dict()
                    for key, col_nr in self.tp_map.items():
                        if key == "channel_type":
                            row_data[key] = channel_type
                        else:
                            try:
                                row_data[key] = row[col_nr].value
                            except:
                                row_data[key] = ""
                    tp_data.append(row_data)

            except:
                self.callback_status(f"ERRROR reading in row nr: {i}")
        self.callback_progress(100)
        self.callback_status("reading data from TP - COMPLETE")
        return tp_data

    def _getTempDirName(self, tempFileName="temp_channel.xlsm") -> os.path:
        """
        creates temp dir if not existing and returns path to dir and filename
        :param tempFileName: Filename for the temporary file
        :return:
        """
        workingdir = os.path.abspath(os.path.dirname(sys.argv[0]))
        TEMP_DIR = os.path.join(workingdir, "temp", "")
        if not os.path.exists(TEMP_DIR):
            os.makedirs(TEMP_DIR)
        TEMP_NAME = os.path.join(TEMP_DIR, tempFileName)
        if os.path.exists(TEMP_NAME):
            os.remove(TEMP_NAME)
        return TEMP_NAME

    def _removePasswordFromExcelFile(self, filepath, pw_str, new_filepath):
        """
        speichert Excel-Datei unter neuem Namen ab und entfernt dabei das Passwort
        :param filepath: path to an Excel File
        :param pw_str: the password used to open the Excel File
        :param new_filepath: the new filepath where the file should be saved to (without password protection)
        :return:
        """
        self.callback_status("removing Password")
        try:
            file = msoffcrypto.OfficeFile(open(filepath, "rb"))
            # Use password
            file.load_key(password=pw_str)
            file.decrypt(open(new_filepath, "wb"))
        except:
            self.callback_status(traceback.format_exc())
