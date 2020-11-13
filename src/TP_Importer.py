import traceback
import os
import sys
from openpyxl import load_workbook, Workbook
import msoffcrypto
from typing import List, Dict


class TP_Importer():
    def __init__(self):
        self.tp_map = {
            "status": 0,
            "did": 1,
            "tnr": 2,
            "titel_local": 3,
            "titel_ov": 6,
            "lg": 7,
            "deal_type": 9,
            "info_link": 11,
            "mandant": 12,
            "vertrieb_physisch": 13,
            "release_actuality": 15,
            "release_origin": 16,
            "quality": 18,
            "production_country": 19,
            "production_year": 20,
            "genre": 21,
            "theatrical_start": 22,
            "theatrical_admissions": 23,
            "runtime": 24,
            "rating": 25,
            "language_local": 26,
            "language_ov": 27,
            "dvd_start": 29,
            "est_start": 46,
            "est_start_4k": 181,
            "est_end": 48,
            "tvod_start": 50,
            "tvod_start_4k": 182,
            "tvod_end": 51,
            "country_de": 42,
            "country_at": 43,
            "country_ch": 44,
            "country_lux": 59,
            "country_lie": 60,
            "right_est": 45,
            "right_tvod": 49,
            "premium_vod_start": 64,
            "premium_vod_end": 65,
            "premium_vod_price_tier": 66,
            "holdback_est_start": 67,
            "holdback_est_end": 68,
            "holdback_tvod_start": 69,
            "holdback_tvod_end": 70,
            "change_reason": 79,
            "pf_status_itunes": 81,
            "pf_status_amazon": 105,
            "pf_status_google": 97,
            "pf_status_microsoft": 88,
            "pf_status_videoload": 84,
            "pf_status_sony": 86,
            "pf_status_ondemand": 99,
            "studio": 108,
            "vendor_id_itunes": 118,
            "vendor_id_google": 133,
            "vendor_id_amazon": 151,
            "vendor_id_microsoft": 129,
            "vendor_id_sky": 130,
            "vendor_id_sony": 131,
            "vendor_id_vodafone": 134,
            "vendor_id_maxdome": 135,
            "vendor_id_ondemand": 136,
            "vendor_id_videoload": 146,
            "vendor_id_wuaki": 147,
            "vendor_id_hollystar": 148,
            "vendor_id_chili": 149,
            "vendor_id_videociety": 116,
            "vendor_id_videobuster": 115,
            "vendor_id_teleclub": 114,
            "vendor_id_cablecom": 113,
            "vendor_id_magenta_at": 112,
            "vendor_id_unitymedia": 111,
            "ov": 144,
            "full_delete": 139,
            "full_delete_4k_amazon": 140,
            "full_delete_poest": 141,
            "isan": 170,
            "imdb_link": 175,
            "pricing_initial_4k_de": 179,
            "pricing_initial_4k_ch": 180,
            "pricing_1streprice_4k_de": 183,
            "pricing_1streprice_4k_ch": 184,
            "pricing_1streprice_4k_start": 185,
            "pricing_initial_hd_de": 190,
            "pricing_initial_sd_de": 191,
            "pricing_1streprice_start": 193,
            "pricing_1streprice_hd": 236,
            "pricing_1streprice_sd": 237,
            "wsp_initial_sd_de_amazon": 201,
            "wsp_initial_hd_de_amazon": 202,
            "wsp_1streprice_sd_de_amazon": 240,
            "wsp_1streprice_hd_de_amazon": 241,
            "wsp_special_amazon_start": 310,
            "wsp_special_amazon_end": 311,
            "wsp_special_amazon_est_sd": 312,
            "wsp_special_amazon_est_hd": 313,
            "wsp_special_amazon_est_4k": 314,
            "wsp_special_amazon_tvod_sd": 315,
            "wsp_special_amazon_tvod_hd": 316,
            "wsp_special_amazon_tvod_4k": 317,
        }

    def callback_status(self, status: str):
        """
        a callback function that reports status updates - should be overwritten when subclassing this class
        :param status: str
        :return:
        """
        print(status)

    def callback_progress(self, progress: int):
        """
        a callback function that reports progress updates - should be overwritten when subclassing this class
        :param progress: int
        :return:
        """
        print(progress)

    def import_TP(self, path) -> List[Dict]:
        """
        imports the Titelplanung xlsm from a given filepath (String or Path)
        :param path: filepath (String or Path)
        :return: a list of dicts (one dict per title)
        """
        self.callback_status('reading data from TP')
        self.callback_progress(0)
        wb = self._load_workbook(path)
        self.callback_progress(20)
        tp_data = self._get_data_from_wb(wb)
        self.callback_progress(100)
        self.callback_status('reading data from TP - COMPLETE')
        return tp_data

    def _load_workbook(self, path) -> Workbook:
        """
        loads an Excel Workbook from the given path
        :param path:
        :return:
        """
        try:
            wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False, keep_links=False)
        except:
            # falls das Excel File passwortgeschützt ist schlägt der erste Versuch fehl;
            # dann muss erst das Passwort entfernt werden
            TEMP_NAME = self._getTempDirName(tempFileName='temp_tp.xlsm')
            self._removePasswordFromExcelFile(path, 'uf', TEMP_NAME)
            self.callback_status('reading data from TP')
            wb = load_workbook(TEMP_NAME, read_only=True, data_only=True)
        return wb

    def _get_data_from_wb(self, workbook: Workbook, first_row: int=8):
        """
        fetches data from the given Excel Workbook instance
        :param workbook:
        :param first_row:
        :return:
        """
        ws = workbook["TP"]
        tp_data = []
        i = first_row
        for row in ws['A' + str(first_row):'QF10000']:
            try:
                i += 1
                if row[self.tp_map.get('tnr')].value and row[self.tp_map.get('status')].value in (
                        "ok", "change", "new"):
                    row_data = dict()
                    for key, col_nr in self.tp_map.items():
                        try:
                            row_data[key] = row[col_nr].value
                        except:
                            row_data[key] = ""
                    tp_data.append(row_data)
            except:
                self.callback_status(f'ERRROR reading in row nr: {i}')
        return tp_data

    def _getTempDirName(self, tempFileName='temp_channel.xlsm') -> os.path:
        """
        creates temp dir if not existing and returns path to dir and filename
        :param tempFileName: Filename for the temporary file
        :return:
        """
        workingdir = os.path.abspath(os.path.dirname(sys.argv[0]))
        TEMP_DIR = os.path.join(workingdir, 'temp', '')
        if not os.path.exists(TEMP_DIR):
            os.makedirs(TEMP_DIR)
        TEMP_NAME = os.path.join(TEMP_DIR, tempFileName)
        if os.path.exists(TEMP_NAME):
            os.remove(TEMP_NAME)
        return TEMP_NAME

    def _removePasswordFromExcelFile(self, filepath, pw_str, new_filepath):
        """
        speichert Excel-Datei unter neuem Namen ab und entfernt dabei das Passwort
        :param filepath: path to an Excel File
        :param pw_str: the password used to open the Excel File
        :param new_filepath: the new filepath where the file should be saved to (without password protection)
        :return:
        """
        self.callback_status('removing Password')
        try:
            file = msoffcrypto.OfficeFile(open(filepath, "rb"))
            # Use password
            file.load_key(password=pw_str)
            file.decrypt(open(new_filepath, 'wb'))
        except:
            self.callback_status(traceback.format_exc())

if __name__ == '__main__':
    tp_data = TP_Importer().import_TP('G:\Listen\TPDD aktuell absolutiert.xlsm')

    print(tp_data)