from openpyxl import Workbook
from typing import List, Dict, Tuple, Set

from src.tpimporter import Excel_Importer


class TVTP_Importer_Channels(Excel_Importer):
    callback_status_signal = None
    callback_progress_signal = None

    # Hier können neue Channels hinzugefügt werden
    CHANNEL_TYPE_TV = {
        "TV_Plan ACNMA": "arthousecnma",
        "TV_Plan HOH": "homeofhorror",
        "TV_Plan Filmtastic": "filmtastic",
        "TV_Plan Cinehearts": "cinehearts",
        "TV_Plan Filmlegenden": "filmlegenden",
    }

    # SPECIFY FIELDNAMES AND CORRESPONDING EXCEL COLUMN NUMBERs HERE
    tv_series_map = {
        "titel_local": 3,
        "titel_ov": 4,
        "vendor_id": 53,
        "vendor_id_amazon": 46,
        "vendor_id_standalone": 92,
        # die folgenden Spalten gibt es noch nicht in der Channel-TP TV
        # "vendor_id_wuaki": 46,
        # "vendor_id_zattoo": 47,
        # "vendor_id_waipu": 49,
        # "vendor_id_magenta_at": 50,
        "so_number": 95,
        "channel_type": 9999,  # wird nicht aus Excel gelesen
    }

    tv_season_map = {
        "status": 0,
        "did": 1,
        "tnr": 2,
        "so_number": 94,
        "titel_local": 5,
        "titel_ov": 6,
        "number": 7,
        "no_of_episodes": 11,
        "deal_type": 16,
        # "mandant": 27,
        "license_start": 17,
        "license_end": 18,
        "genre": 41,
        "rating": 28,
        "licensor": 14,
        "production_year": 33,
        "production_country": 32,
        "quality": 27,
        "tv_network": 37,
        "studio": 50,
        "vendor_id": 52,
        "vendor_id_amazon": 45,
        "vendor_id_standalone": 91,
        # die folgenden Spalten gibt es noch nicht in der Channel-TP TV
        # "vendor_id_wuaki": 46,
        # "vendor_id_zattoo": 47,
        # "vendor_id_waipu": 49,
        # "vendor_id_magenta_at": 50,
        "pf_status_amazon": 56,
        "pf_status_standalone": 57,
        "pf_status_magenta_at": 58,
        "pf_status_rakuten": 59,
        # die folgenden Spalten gibt es noch nicht in der Channel-TP TV
        # "pf_status_waipu": 90,
        # "pf_status_zattoo": 91,
        "country_de": 24,
        "country_at": 25,
        "country_ch": 26,
        "channel_type": 9999,  # wird nicht aus Excel gelesen
    }

    tv_episode_map = {
        "did": 1,
        "titel_local": 9,
        "titel_ov": 10,
        "number": 12,
        "rating": 28,
        "license_start": 17,
        "license_end": 18,
        "tv_air_date": 36,
        "runtime": 13,
        "vendor_id": 51,
        "vendor_id_amazon": 44,
        "vendor_id_standalone": 90,
        # die folgenden Spalten gibt es noch nicht in der Channel-TP TV
        # "vendor_id_wuaki": 46,
        # "vendor_id_zattoo": 47,
        # "vendor_id_waipu": 49,
        # "vendor_id_magenta_at": 50,
        "so_number": 93,
        "channel_type": 9999,  # wird nicht aus Excel gelesen
    }

    def __init__(self, valid_statuses=("ok", "change", "new")):
        super().__init__(valid_statuses)

    def get_tp_data_from_file(self, path) -> Dict:
        """
        imports the Channel-Titelplanung (TV) xlsm from a given filepath (String or Path)
        :param path: filepath (String or Path)
        :return: a dict with each channel-type as a key and a dictionary as value -
        (one entry per series with the series-Basis-VendorID as key) - die Daten zu den Staffeln
        kann man als Liste von Dicts unter dem Key "seasons" aufrufen
        (auch wieder als Dict mit Key = Season-Basis-VID),
        die zu den Episoden sind im Staffel-Dict als Liste von Dicts unter dem Key "episodes" zu finden
        (auch wieder als Dict mit Key = Episode-Basis-VID)
        """
        wb = self._load_workbook(path)
        channel_tp_data_tv = dict()
        for worksheet_name, channel_type in self.CHANNEL_TYPE_TV.items():
            tp_data = self._get_data_from_wb(
                workbook=wb,
                first_row=4,
                worksheet_name=worksheet_name,
                channel_type=channel_type,
            )
            channel_tp_data_tv[channel_type] = tp_data
        return channel_tp_data_tv

    def _get_data_from_wb(
        self,
        workbook: Workbook,
        first_row: int = 4,
        worksheet_name: str = "TPTV",
        channel_type="",
    ) -> Dict:
        """
        fetches data from the given Excel Workbook instance
        :param workbook: a Workbook instance
        :param first_row: the first row where we should expect the data to begin
        :param worksheet_name: name of the worksheet that should be read
        :param channel_type: possible values: transactional/filmtastic/arthousecnma/homeofhorror/cinehearts/filmlegenden
        :return: a dictionary of all TV_Series --> every series is a dictionary that contains series info
        and under the key 'seasons' you'll find multiple dictionaries for all seasons of the series; likewise in every
        season you'll find the episodes under the key 'episodes'
        """
        series_data = dict()
        ws = workbook[worksheet_name]
        i = first_row
        max_row = ws.max_row
        print("max_row:", max_row)
        for row in ws["A" + str(first_row) : "HB" + str(max_row)]:
            try:
                # progress (20 % schon nach öffnen erreicht)
                self.callback_progress(20 + int(i / max_row * 80))

                # nur valide Items mit aktivem Status und einer Titelnummer nehmen
                if (
                    row[self.tv_season_map.get("tnr")].value
                    and row[self.tv_season_map.get("status")].value
                    in self.valid_statuses
                ):
                    # SERIES
                    series_dict = self._get_row_data(
                        row=row, item_map=self.tv_series_map, channel_type=channel_type
                    )
                    if not series_dict.get("vendor_id") in series_data.keys():
                        series_data[series_dict.get("vendor_id")] = series_dict

                    # SEASONS
                    if str(row[8].value).lower().strip() == "x":
                        season_data = self._get_row_data(
                            row=row,
                            item_map=self.tv_season_map,
                            channel_type=channel_type,
                        )
                        series = series_data.get(
                            row[self.tv_series_map.get("vendor_id")].value
                        )
                        if not "seasons" in series.keys():
                            series["seasons"] = {
                                season_data.get("vendor_id"): season_data
                            }
                        else:
                            series["seasons"][
                                season_data.get("vendor_id")
                            ] = season_data

                    # EPISODES
                    episode_data = self._get_row_data(
                        row=row, item_map=self.tv_episode_map, channel_type=channel_type
                    )
                    series = series_data.get(
                        row[self.tv_series_map.get("vendor_id")].value
                    )
                    season = series.get("seasons").get(
                        row[self.tv_season_map.get("vendor_id")].value
                    )
                    if not "episodes" in season.keys():
                        season["episodes"] = {
                            episode_data.get("vendor_id"): episode_data
                        }
                    else:
                        season["episodes"][episode_data.get("vendor_id")] = episode_data

                i += 1

            except:
                self.callback_status(f"ERRROR reading in row nr: {i}")

        self.callback_progress(100)
        self.callback_status("reading data from TP - COMPLETE")
        return series_data

    def _get_row_data(self, row, item_map: dict, channel_type: str) -> dict:
        """
        reads in data from an Excel-row; specify the data you want within the item_map dictionary
        (field_name:Column-Number in Excel);
        :param row: an Excel row
        :param item_map: a dictionary (field_name:Column-Number in Excel)
        :param channel_type: String; possible values: transactional/filmtastic/arthousecnma/homeofhorror
        :return: dictionary with the keys from item_map and the corresponding values from the Excel row
        """
        row_data = dict()
        for key, col_nr in item_map.items():
            if key == "channel_type":
                row_data[key] = channel_type
            else:
                try:
                    row_data[key] = row[col_nr].value
                    if row[col_nr].value == "#N/A":
                        row_data[key] = None
                except:
                    row_data[key] = ""
        return row_data


if __name__ == "__main__":
    channel_tp_data_feature = dict()
    importer = TVTP_Importer_Channels(valid_statuses=(("ok", "change", "new")))
    path = "G:\Listen\Titelplanung Channels aktuell_absolutiert_new.xlsm"
    tp_data = importer.get_tp_data_from_file(path)
    print(tp_data.keys())
    for channel_type, channel_titles in tp_data.items():
        # durch alle Channels iterieren
        for vendor_id, series in channel_titles.items():
            # do something for series
            print(series)
            for vendor_id_season, season in series.get("seasons", dict()).items():
                # do something for seasons
                print(season)
                for vendor_id_episode, episode in season.get(
                    "episodes", dict()
                ).items():
                    # do something for episodes
                    print(episode)
