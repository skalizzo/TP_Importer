from openpyxl import Workbook
from typing import List, Dict
from src.tpimporter import Excel_Importer


class TP_Importer_Channels(Excel_Importer):
    callback_status_signal = None
    callback_progress_signal = None

    CHANNEL_TYPE_FEATURE = {
        "Planung ACNMA": "arthousecnma",
        "Planung HOH": "homeofhorror",
        "Planung Filmtastic": "filmtastic",
        "Planung Cinehearts": "cinehearts",
        "Planung Filmlegenden": "filmlegenden",
    }

    tp_map = {
        "status": 1,
        "did": 2,
        "tnr": 3,
        "titel_local": 4,
        "titel_ov": 5,
        "lg": 8,
        "quality": 21,
        "production_country": 27,
        "production_year": 28,
        "genre": 26,
        "theatrical_start": 32,
        "theatrical_admissions": 33,
        "dvd_start": 31,
        "license_start": 11,
        "license_end": 12,
        "country_de": 18,
        "country_at": 19,
        "country_ch": 20,
        # "right_svod": 52,
        # 'mandant'
        "pf_status_magenta_at": 88,
        "pf_status_rakuten": 89,
        "pf_status_waipu": 90,
        "pf_status_zattoo": 91,
        "pf_status_amazon": 86,
        "pf_status_standalone": 87,
        "studio": 44,
        "vendor_id": 0,
        "vendor_id_amazon": 40,
        "vendor_id_wuaki": 46,
        "vendor_id_zattoo": 47,
        "vendor_id_standalone": 48,
        "vendor_id_waipu": 49,
        "vendor_id_magenta_at": 50,
        "channel_category": 55,
        "imdb_link": 36,
        "so_number": 59,
    }

    def __init__(self, valid_statuses=("ok", "change", "new")):
        super().__init__(valid_statuses)

    def get_tp_data_from_file(self, path) -> Dict[str, List[dict]]:
        """
        imports the Channel-Titelplanung xlsm from a given filepath (String or Path)
        :param path: filepath (String or Path)
        :return: a dict with channel types as keys and a list of dictionaries as value
        (one dict for every title within the channel)
        """
        wb = self._load_workbook(path)
        channel_tp_data_feature = dict()
        for worksheet_name, channel_type in self.CHANNEL_TYPE_FEATURE.items():
            tp_data = self._get_data_from_wb(
                workbook=wb,
                first_row=4,
                worksheet_name=worksheet_name,
                channel_type=channel_type,
            )
            channel_tp_data_feature[channel_type] = tp_data
        return channel_tp_data_feature

    def _get_data_from_wb(
        self,
        workbook: Workbook,
        first_row: int = 4,
        worksheet_name: str = "Planung_ACNMA",
        channel_type="arthousecnma",
    ) -> List[Dict]:
        """
        fetches data from the given Excel Workbook instance
        :param workbook: a Workbook instance
        :param first_row: the first row where we should expect the data to begin
        :param worksheet_name: name of the worksheet that should be read
        :param channel_type: possible values: transactional/filmtastic/arthousecnma/homeofhorror/cinehearts/filmlegenden
        :return:
        """
        tp_data = []
        ws = workbook[worksheet_name]
        i = first_row
        max_row = ws.max_row
        print("max_row:", max_row)
        for row in ws["A" + str(first_row) : "QF" + str(max_row)]:
            try:
                # progress (20 % schon nach öffnen erreicht)
                self.callback_progress(20 + int(i / max_row * 80))
                i += 1
                if (
                    row[self.tp_map.get("tnr")].value
                    and row[self.tp_map.get("status")].value in self.valid_statuses
                ):
                    row_data = dict()
                    for key, col_nr in self.tp_map.items():
                        if key == "channel_type":
                            row_data[key] = channel_type
                        else:
                            try:
                                if type(row[col_nr].value) == str:
                                    row_data[key] = str(row[col_nr].value).strip()
                                else:
                                    row_data[key] = row[col_nr].value
                            except:
                                row_data[key] = ""
                    tp_data.append(row_data)

            except:
                self.callback_status(f"ERRROR reading in row nr: {i}")
        self.callback_progress(100)
        self.callback_status("reading data from TP_Channels - COMPLETE")
        return tp_data


if __name__ == "__main__":
    channel_tp_data_feature = dict()
    importer = TP_Importer_Channels(valid_statuses=(("ok", "change", "new")))
    path = "G:\Listen\Titelplanung Channels aktuell_absolutiert_new.xlsm"
    tp_data = importer.get_tp_data_from_file(path)
    print(tp_data.keys())
